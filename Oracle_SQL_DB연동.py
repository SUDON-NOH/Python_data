# oracle_bs4.py

import requests
import urllib
import cx_Oracle

import pandas as pd
import chardet
from openpyxl import Workbook
# pip install openpyxl

# CREATE TABLE COSMATIC  (
#     BRAND    VARCHAR2(100),
#     NAME     VARCHAR2(100),
#     PRICE    VARCHAR2(100),
#     SRC      VARCHAR2(1000)
# );
#
# DROP TABLE COSMATIC;
#
# DESC COSMATIC;

from bs4 import BeautifulSoup

#####한글깨짐 방지######
import os
os.environ["NLS_LANG"] = ".AL32UTF8"

# # Workbook을 통해 엑셀 파일을 생성
# wb = Workbook()
#
# # 현재 Active Sheet 얻기
# ws = wb.active
# ws.title = "first_Steet"
# # 에겔 컬럼에 따른 컬럼명
# ws["A1"] = "brand"
# ws["B1"] = "name"
# ws["C1"] = "price"
# ws["D1"] = "src"

# DB와 연결된 코드
conn = cx_Oracle.connect('scott/tiger@localhost:1521/orcl')

print(conn.version)

cursor = conn.cursor()
cursor.execute("""CREATE TABLE COSMATIC  (
    BRAND    VARCHAR2(100),
    NAME     VARCHAR2(30),
    PRICE    VARCHAR2(20),
    SRC      VARCHAR2(1000) 
)""")

# 어떤 특정 웹 페이지를 크롤할 수 있게 하는 함수정의입니다. parameter는 최대 클롤할 페이지 수를 받습니다.
def spider(max_pages):
    # page = 1 초기 입력해 주는 페이지가 최대 크롤할 페이지보다 작을 동안 계속 크롤할 수 있게 while 반복문 사용
    page = 1
    urlpage = 0
    i = 2

    while page <= max_pages:

        # 크롤할 사이트를 본 블로그로 하고 그 뒤에 페이지 번호를 합쳐 저장한 값을 변수 url에 저장하도록 하였습니다.
        url1 = 'https://www.yslbeautykr.com/ko_KR/makeup/lips?sz=12&start='
        # url1 = 'https://www.yslbeautykr.com/ko_KR/makeup/lips/lipstick'
        url = url1 + str(urlpage)

        #  source_code = requests.get(url) 타겟 url로 부터 데이터를 저장할 수 있다.
        source_code = requests.get(url)

        #  가져온 데이터를 source_code의 텍스트 부분만을 처리하여 plain_text에 저장
        plain_text = source_code.text

        # 여기서 실행을 시키게 되면 무슨 내용이 있는지 잘 알수 가 없는 데이터로 가져오게 된다.
        # 그래서, 아래와 같이 데이터를 사람이 일기 쉽게 해주는 파이썬 모듈이 초기에 불러온다.
        # Beautifulsoup이라는 모듈이다.
        # 가독성을 높이기 위해 Beautifulsoup(타겟, parser)를 처리하여 임의의 변수인 soup에 저장
        soup = BeautifulSoup(plain_text, 'html.parser')

        for div in soup.findAll('div', {'class': 'product_tile'}):

            p_tag = div.find('a', {'class': 'product_name desktop_content tablet_content'}).getText()
            name_tag = div.find('div', {'class': 'product_subtitle'}).getText()
            span_tag = div.find('p',
                                {'class': 'product_price price_sale b-product_price-sale b-product_price'}).getText()
            img_tag = div.find('img')
            print(img_tag)

            print(p_tag.strip())
            print(name_tag.strip())
            print(span_tag.strip())
            print(img_tag.get('data-desktop-src').strip())



            # nametemp = img_tag.get('alt')
            # if len(nametemp) == 0:
            #     filename = str(i)
            #     i = i + 1
            # else:
            #     filename = nametemp
            #
            # imagesfile = open(filename + ".jpeg", 'wb')
            # imagesfile.write(urllib.request.urlopen(img_tag.get('data-desktop-src')).read())
            # imagesfile.close()
            #
            # ws.cell(row=i, column=1, value=p_tag.strip())
            # ws.cell(row=i, column=2, value=name_tag.strip())
            # ws.cell(row=i, column=3, value=span_tag.strip())
            # ws.cell(row=i, column=4, value=img_tag.get('data-desktop-src').strip())

            # insert
            sql_insert = 'insert into COSMATIC VALUES(:BRAND, :NAME, :PRICE, :SRC)'
            db = conn.cursor()
            # encode 디비에 들어가기는 하나... 이상한 값이 출력됨
            # encode()로 변환한다음 decode()로 한글 변환한다.
            db.execute(sql_insert,
                       BRAND=p_tag.encode('utf8').decode('utf8'),
                       NAME=name_tag.strip().encode('utf8').decode('utf8'),
                       PRICE=span_tag.strip().encode('utf8').decode('utf8'),
                       SRC =img_tag.get('data-desktop-src').encode('utf8').decode('utf8') )

            conn.commit()

            db.execute('SELECT * FROM COSMATIC')

            for record in db:
                print(record)

            i = i + 1

        # while반복문이 무한반복이 되지 않게 하기 위해 page += 1을 주어 초기 페이지부터 지정한 특정페이지 번위의 모든 타이틀과 url값을 출력할 수 있다.
        page += 1
        urlpage += 12


    cursor.execute("SELECT * FROM COSMATIC")
    row = cursor.fetchall()
    header = ['BRAND', 'NAME', 'PRICE', 'SRC' ]

    df = pd.DataFrame(row, columns=header)
    print(df)

    df.to_excel("cosmatic.xlsx")

    cursor.execute("""DROP TABLE COSMATIC""")

    db.close()
    conn.close()

# 1페이지만 크롤링된다. 만약 spider(!0)이라면 1~9까지 스크롤하게 max_page값을 변경하게됩니다.
spider(2)


# wb.save("Yves Saint Laurent.xlsx")
